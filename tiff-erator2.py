import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import fitz  # PyMuPDF
from PIL import Image, ImageDraw, ImageFont, ImageOps
from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlrd
import pandas as pd
import chardet
import extract_msg
import email
from email import policy
from email.parser import BytesParser

# Global variable to control the conversion process
stop_conversion = False

def convert_eml_to_tiff(eml_path, tiff_path, dpi=150, compression="tiff_lzw"):
    with open(eml_path, 'rb') as f:
        msg = BytesParser(policy=policy.default).parse(f)

    # Try to get the plain text body
    msg_body = msg.get_body(preferencelist=('plain'))
    if msg_body is None:
        # If plain text body is not available, try to get the HTML body
        msg_body = msg.get_body(preferencelist=('html'))
        if msg_body is not None:
            msg_body = msg_body.get_content()
        else:
            # If neither plain text nor HTML body is available, set a default message
            msg_body = "No content available"
    else:
        msg_body = msg_body.get_content()

    # Create a blank image with white background
    img = Image.new("RGB", (2480, 3508), (255, 255, 255))  # A4 size at 300 DPI
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 12)
    except IOError:
        font = ImageFont.load_default()

    # Draw the email body text onto the image
    d.text((10, 10), msg_body, fill=(0, 0, 0), font=font)

    # Save the image as a TIFF file
    img.save(tiff_path, compression=compression)

def convert_msg_to_tiff(msg_path, tiff_path, dpi=150, compression="tiff_lzw"):
    msg = extract_msg.Message(msg_path)
    msg_body = msg.body

    # Create a blank image with white background
    img = Image.new("RGB", (2480, 3508), (255, 255, 255))  # A4 size at 300 DPI
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 12)
    except IOError:
        font = ImageFont.load_default()

    # Draw the email body text onto the image
    d.text((10, 10), msg_body, fill=(0, 0, 0), font=font)

    # Save the image as a TIFF file
    img.save(tiff_path, compression=compression)

def convert_files(input_folder, output_folder, dpi=150, compression="tiff_lzw"):
    global stop_conversion
    stop_conversion = False
    files = os.listdir(input_folder)
    total_files = len(files)
    for index, filename in enumerate(files):
        if stop_conversion:
            messagebox.showinfo("Stopped", "Conversion process has been stopped.")
            return
        file_path = os.path.join(input_folder, filename)
        tiff_path = os.path.join(output_folder, os.path.splitext(filename)[0] + ".tiff")
        if filename.endswith(".pdf"):
            convert_pdf_to_tiff(file_path, tiff_path, dpi, compression)
        elif filename.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif")):
            convert_image_to_tiff(file_path, tiff_path, compression)
        elif filename.endswith(".docx"):
            convert_docx_to_tiff(file_path, tiff_path, dpi, compression)
        elif filename.lower().endswith((".xlsx", ".xls", ".csv")):
            convert_spreadsheet_to_tiff(file_path, tiff_path, dpi, compression)
        elif filename.endswith(".msg"):
            convert_msg_to_tiff(file_path, tiff_path, dpi, compression)
        elif filename.endswith(".eml"):
            convert_eml_to_tiff(file_path, tiff_path, dpi, compression)
        counter_label.config(text=f"Converting {index + 1} out of {total_files} files")
    messagebox.showinfo("Success", "All files have been converted to TIFF.")

def convert_pdf_to_tiff(pdf_path, tiff_path, dpi=150, compression="tiff_lzw"):
    doc = fitz.open(pdf_path)
    images = []
    for page_num in range(len(doc)):
        if stop_conversion:
            return
        page = doc.load_page(page_num)
        zoom = dpi / 72  # 72 is the default DPI for PDFs
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(img)

    if images:
        images[0].save(tiff_path, save_all=True, append_images=images[1:], compression=compression)

def convert_image_to_tiff(image_path, tiff_path, compression="tiff_lzw"):
    img = Image.open(image_path)
    img.save(tiff_path, compression=compression)

def convert_docx_to_tiff(docx_path, tiff_path, dpi=150, compression="tiff_lzw"):
    doc = Document(docx_path)
    images = []
    for para in doc.paragraphs:
        if stop_conversion:
            return
        img = Image.new("RGB", (2480, 3508), (255, 255, 255))  # A4 size at 300 DPI
        d = ImageDraw.Draw(img)
        d.text((10, 10), para.text, fill=(0, 0, 0))
        images.append(img)
    if images:
        images[0].save(tiff_path, save_all=True, append_images=images[1:], compression=compression)

def is_cell_populated(cell):
    return cell.value is not None or cell.has_style

def set_print_area_if_needed(ws):
    min_row, max_row, min_col, max_col = None, None, None, None
    for row in ws.iter_rows():
        for cell in row:
            if is_cell_populated(cell):
                if min_row is None or cell.row < min_row:
                    min_row = cell.row
                if max_row is None or cell.row > max_row:
                    max_row = cell.row
                if min_col is None or cell.column < min_col:
                    min_col = cell.column
                if max_col is None or cell.column > max_col:
                    max_col = cell.column
    if min_row and max_row and min_col and max_col:
        ws.print_area = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

def count_print_pages(ws, orientation, file_type):
    if file_type == 'xlsx':
        ws.page_setup.orientation = orientation
        set_print_area_if_needed(ws)
        if isinstance(ws.print_area, str):
            row_count = int(ws.print_area.split(':')[1][1:])
            col_count = ord(ws.print_area.split(':')[1][0].upper()) - 64
        else:
            row_count = ws.max_row
            col_count = ws.max_column
    elif file_type == 'xls':
        row_count = ws.nrows
        col_count = ws.ncols
    else:  # csv
        row_count = len(ws)
        col_count = len(ws.columns)

    rows_per_page = 50 if orientation == 'portrait' else 30
    cols_per_page = 10 if orientation == 'portrait' else 15

    row_pages = (row_count // rows_per_page) + 1
    col_pages = (col_count // cols_per_page) + 1

    return row_pages * col_pages

def detect_encoding(file_path):
    with open(file_path, 'rb') as file:
        raw_data = file.read()
    return chardet.detect(raw_data)['encoding']

def read_file_with_encoding(file_path):
    _, ext = os.path.splitext(file_path.lower())

    if ext == '.xlsx':
        return pd.read_excel(file_path)
    elif ext == '.xls':
        return pd.read_excel(file_path, engine='xlrd')
    elif ext == '.csv':
        encoding = detect_encoding(file_path)
        return pd.read_csv(file_path, encoding=encoding)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

def set_optimal_orientation(file_path):
    _, ext = os.path.splitext(file_path.lower())

    try:
        if ext == '.xlsx':
            wb = load_workbook(file_path)
            ws = wb.active
            file_type = 'xlsx'
        elif ext == '.xls':
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            file_type = 'xls'
        elif ext == '.csv':
            df = read_file_with_encoding(file_path)
            ws = df
            file_type = 'csv'
        else:
            raise ValueError(f"Unsupported file type: {ext}")

        portrait_pages = count_print_pages(ws, 'portrait', file_type)
        landscape_pages = count_print_pages(ws, 'landscape', file_type)
        optimal_orientation = 'landscape' if landscape_pages < portrait_pages else 'portrait'

        if file_type == 'xlsx':
            ws.page_setup.orientation = optimal_orientation
            wb.save(file_path)

        return optimal_orientation
    except Exception as e:
        print(f"Error in set_optimal_orientation: {str(e)}")
        return 'portrait'  # Default to portrait if there's an error

def convert_spreadsheet_to_tiff(file_path, tiff_path, dpi=150, compression="tiff_lzw"):
    try:
        _, ext = os.path.splitext(file_path.lower())
        images = []
        if ext == '.xlsx':
            wb = load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.page_setup.fitToWidth = 1  # Fit all columns on one page
                optimal_orientation = set_optimal_orientation(file_path)
                df = pd.DataFrame(ws.values)
                img_width = 3508 if optimal_orientation == 'landscape' else 2480
                img_height = 2480 if optimal_orientation == 'landscape' else 3508
                img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
                d = ImageDraw.Draw(img)
                try:
                    font = ImageFont.truetype("arial.ttf", 12)
                except IOError:
                    font = ImageFont.load_default()
                y_offset = 10
                for _, row in df.iterrows():
                    x_offset = 10
                    for item in row:
                        d.text((x_offset, y_offset), str(item), fill=(0, 0, 0), font=font)
                        x_offset += 100
                    y_offset += 20
                    if y_offset >= img_height - 20:
                        images.append(img)
                        img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
                        d = ImageDraw.Draw(img)
                        y_offset = 10
                images.append(img)
        elif ext == '.xls':
            wb = xlrd.open_workbook(file_path)
            for sheet_index in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_index)
                df = pd.DataFrame([sheet.row_values(row) for row in range(sheet.nrows)])
                img_width = 3508
                img_height = 2480
                img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
                d = ImageDraw.Draw(img)
                try:
                    font = ImageFont.truetype("arial.ttf", 12)
                except IOError:
                    font = ImageFont.load_default()
                y_offset = 10
                for _, row in df.iterrows():
                    x_offset = 10
                    for item in row:
                        d.text((x_offset, y_offset), str(item), fill=(0, 0, 0), font=font)
                        x_offset += 100
                    y_offset += 20
                    if y_offset >= img_height - 20:
                        images.append(img)
                        img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
                        d = ImageDraw.Draw(img)
                        y_offset = 10
                images.append(img)
        elif ext == '.csv':
            df = read_file_with_encoding(file_path)
            img_width = 3508
            img_height = 2480
            img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
            d = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 12)
            except IOError:
                font = ImageFont.load_default()
            y_offset = 10
            for _, row in df.iterrows():
                x_offset = 10
                for item in row:
                    d.text((x_offset, y_offset), str(item), fill=(0, 0, 0), font=font)
                    x_offset += 100
                y_offset += 20
                if y_offset >= img_height - 20:
                    images.append(img)
                    img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
                    d = ImageDraw.Draw(img)
                    y_offset = 10
            images.append(img)
        if images:
            images[0].save(tiff_path, save_all=True, append_images=images[1:], compression=compression)
    except Exception as e:
        print(f"Error in convert_spreadsheet_to_tiff: {str(e)}")

def select_input_folder():
    folder_selected = filedialog.askdirectory()
    input_folder_var.set(folder_selected)

def select_output_folder():
    folder_selected = filedialog.askdirectory()
    output_folder_var.set(folder_selected)

def start_conversion():
    input_folder = input_folder_var.get()
    output_folder = output_folder_var.get()
    if not input_folder or not output_folder:
        messagebox.showwarning("Input Error", "Please select both input and output folders.")
        return
    conversion_thread = threading.Thread(target=convert_files, args=(input_folder, output_folder))
    conversion_thread.start()

def stop_conversion_process():
    global stop_conversion
    stop_conversion = True

root = tk.Tk()
root.title("Tiff-erator")

input_folder_var = tk.StringVar()
output_folder_var = tk.StringVar()

tk.Label(root, text="Select Input Folder:").pack(pady=5)
tk.Entry(root, textvariable=input_folder_var, width=50).pack(pady=5)
tk.Button(root, text="Browse", command=select_input_folder).pack(pady=5)

tk.Label(root, text="Select Output Folder:").pack(pady=5)
tk.Entry(root, textvariable=output_folder_var, width=50).pack(pady=5)
tk.Button(root, text="Browse", command=select_output_folder).pack(pady=5)

tk.Button(root, text="Start Conversion", command=start_conversion).pack(pady=20)
tk.Button(root, text="Stop Conversion", command=stop_conversion_process).pack(pady=5)

counter_label = tk.Label(root, text="Converted 0 out of 0 files")
counter_label.pack(pady=10)

root.mainloop()
